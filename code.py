import openpyxl as xl
wb = xl.load_workbook('<file_Name.xlsx>', data_only=True)
sheet = wb['<column1 >']
title_value= sheet.cell(2, 8)
title_value.value = '<column2 >'
title_value= sheet.cell(2, 9)
title_value.value = '<column3 >'
title_value= sheet.cell(2, 10)
title_value.value = '<column3 >'
title_value= sheet.cell(2, 11)
title_value.value = '<column4 >'
for row in range(3, sheet.max_row + 1):
    cell_cat = sheet.cell(row, 11)
    cell_cat.value = f'=VLOOKUP(A{row},Code!B:E,4,FALSE)'
    cell_NP = sheet.cell(row, 8)
    cell_NP.value = f'=VLOOKUP(A{row},Code!B:D,3,FALSE)'
    cell_stock = sheet.cell(row, 7)
    cell_sales = sheet.cell(row, 6)
    cell_stock_new = sheet.cell(row, 9)
    cell_sales_new = sheet.cell(row, 10)
    cell_stock_new.value = f'=VLOOKUP(A{row},Code!B:D,3,FALSE)*{cell_stock.value}'
    cell_sales_new.value = f'=VLOOKUP(A{row},Code!B:D,3,FALSE)*{cell_sales.value}'
wb.save('Led sales Followup.xlsx')
